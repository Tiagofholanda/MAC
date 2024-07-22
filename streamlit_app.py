import tkinter as tk
from tkinter import filedialog, messagebox, ttk, font
import pandas as pd
import os
from datetime import datetime
import win32com.client as win32
from PIL import ImageTk, Image
from fpdf import FPDF


# Função para verificar e instalar as dependências necessárias
def verificar_instalar_dependencias():
    import subprocess
    import sys

    try:
        import pandas
        import openpyxl
        import win32com.client
        import xlsxwriter
        import PIL
        import fpdf
    except ImportError:
        print("Instalando dependências necessárias...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", "requirements.txt"])
        print("Dependências instaladas.")

# Verificar e instalar dependências
verificar_instalar_dependencias()

#DEFININDO FUNÇÕES DE ANALISE:
# Funções para análise de Preenchimento, Tipos, Ortografia, Nomes, Intervalos, Condicionais

def AnalisePreenchimento(df_bd, df_matriz):
    resultado = []

    for coluna in range(1, len(df_matriz.columns)):
        texto = df_matriz.columns[coluna]
        regra = df_matriz.iloc[4, coluna]

        if texto in df_bd.columns:
            for linha in range(len(df_bd)):
                valor_celula = df_bd.loc[linha, texto]

                if (regra == "OBRIGATÓRIO" and pd.isna(valor_celula)) or (regra == "VAZIO" and not pd.isna(valor_celula)):
                    motivo = "Preenchimento obrigatório" if regra == "OBRIGATÓRIO" else "Sem preenchimento"
                    resultado.append({
                        'Data': pd.Timestamp.now(),
                        'Linha': linha+1,
                        'Coluna': texto,
                        'Valor': valor_celula,
                        'Analise': 'Preenchimento'
                    })

    return pd.DataFrame(resultado)

def AnalisarTipos(df_bd, df_matriz):
    resultados = []
    
    for coluna in range(2, len(df_matriz.columns)):
        texto = df_matriz.columns[coluna] 
        parametro = df_matriz.iloc[5, coluna]  
        
        if texto in df_bd.columns:
            if parametro == "TEXTO ABC":
                filtro_alfabetico = df_bd[texto].apply(lambda x: str(x).replace(" ", "").isalpha())
                inconsistencias = df_bd[~filtro_alfabetico]
            elif parametro == "DATA":
                filtro_data_invalida = pd.to_datetime(df_bd[texto], errors='coerce').isna()
                inconsistencias = df_bd[filtro_data_invalida]
            elif parametro == "NUMERO":
                filtro_nao_numerico = pd.to_numeric(df_bd[texto], errors='coerce').isna()
                inconsistencias = df_bd[filtro_nao_numerico]
            else:
                inconsistencias = pd.DataFrame()

            for index, row in inconsistencias.iterrows():
                resultados.append({
                    'Data': pd.Timestamp.now(),
                    'Linha':index+1,
                    'Coluna': texto,
                    'Valor': row[texto],
                    'Analise': 'Tipo'
                })
    
    return pd.DataFrame(resultados)

def main_verificar_ortografia(df_matriz, df_bd):
    erros_ortograficos = []

    for coluna in df_matriz.iloc[6].dropna().index:
        if df_matriz.iloc[6, df_matriz.columns.get_loc(coluna)] == "DICIONÁRIO ORTOGRÁFICO":
            erros_ortograficos += verificar_ortografia(coluna, df_bd)

    return pd.DataFrame(erros_ortograficos)

def verificar_ortografia(nome_coluna, df_bd):
    WordApp = iniciar_aplicativo_word()
    erros_ortograficos = []

    try:
        for indice, valor in df_bd[nome_coluna].items():
            if pd.notnull(valor):
                if not verificar_ortografia_word(valor, WordApp):
                    erros_ortograficos.append({
                        'Data': datetime.now(),
                        'Linha': indice+1,
                        'Coluna': nome_coluna,
                        'Valor': valor,
                        'Analise': 'Ortografia'
                    })
    finally:
        fechar_aplicativo_word(WordApp)

    return erros_ortograficos

def iniciar_aplicativo_word():
    WordApp = win32.Dispatch("Word.Application")
    WordApp.Visible = False
    return WordApp

def fechar_aplicativo_word(WordApp):
    if WordApp is not None:
        WordApp.Quit()

def verificar_ortografia_word(palavra, WordApp):
    try:
        return WordApp.CheckSpelling(palavra)
    except Exception as e:
        print(f"Erro ao verificar ortografia: {e}")
        return True

def verificar_nomes(df_matriz, df_bd, df_nomes):
    erros_encontrados = []

    for coluna in df_matriz.iloc[7].dropna().index:
        if df_matriz.iloc[7, df_matriz.columns.get_loc(coluna)] == "BANCO DE NOMES":
            for indice, valor in enumerate(df_bd[coluna]):
                if pd.notnull(valor):
                    partes_nome = valor.split()
                    for parte_nome in partes_nome:
                        if not nome_eh_valido(parte_nome, df_nomes):
                            erros_encontrados.append({
                                'Data': datetime.now(),
                                'Linha': indice + 1,
                                'Coluna': coluna,
                                'Valor': valor,
                                'Analise': 'Nomes'
                            })

    return pd.DataFrame(erros_encontrados)

def nome_eh_valido(nome, df_nomes):
    return nome in df_nomes.iloc[:, 0].values

def verificar_intervalos(df_matriz, df_bd, df_intervalos):
    df_bd1 = df_bd.fillna(value='')
    inconsistencias = []
    row = df_matriz.iloc[8]

    for col in df_matriz.columns:
        if 'intervalo' in str(row[col]):
            intervalo = str(row[col])
            if col in df_bd1.columns:
                for idx, valor_bd in enumerate(df_bd1[col]):
                    try:
                        float(valor_bd)
                        if valor_bd not in df_intervalos[intervalo]:
                            hora_analise = datetime.now()
                            inconsistencias.append({
                                'Data': hora_analise,
                                'Linha': idx + 1,
                                'Coluna': col,
                                'Valor': valor_bd,
                                'Analise': 'Intervalos'
                            })
                    except:
                        if str(valor_bd) not in str(df_intervalos[intervalo]):
                            hora_analise = datetime.now()
                            inconsistencias.append({
                                'Data': hora_analise,
                                'Linha': idx + 1,
                                'Coluna': col,
                                'Valor': valor_bd,
                                'Analise': 'Intervalos'
                            })

    return pd.DataFrame(inconsistencias)


def verificar_condicionais(df_matriz, df_bd):
    inconsistencias = []
    df_matriz1 = df_matriz.fillna(value='')

    for linha in range(15, 92, 4):
        linha_matriz = df_matriz1.iloc[linha]
        for col in range(3, len(linha_matriz)):
            valor = linha_matriz.iloc[col]
            if valor != "":
                coluna_condicionante = valor
                coluna_condicionada = df_matriz.columns[col]
                dado_condicionante = df_matriz1.iloc[linha + 1, col]
                resultado = df_matriz1.iloc[linha + 2, col]
                # Verifica se a coluna condicionante existe em df_bd
                if coluna_condicionante in df_bd.columns:
                    for indice, valor_bd in enumerate(df_bd[coluna_condicionante]):
                        if valor_bd == dado_condicionante:
                            if coluna_condicionada in df_bd.columns:
                                if df_bd[coluna_condicionada][indice] != resultado:
                                    inconsistencias.append({
                                        'Data': datetime.now(),
                                        'Linha': indice + 1,
                                        'Coluna': coluna_condicionada,
                                        'Valor': valor_bd,
                                        'Analise': 'Condicionais'
                                    })
                else:
                    print(f'Coluna condicionante "{coluna_condicionante}" não encontrada em df_bd.')

    return pd.DataFrame(inconsistencias)

def exportar_resultados_excel():
    global resultado_concatenado

    if resultado_concatenado is not None and not resultado_concatenado.empty:
        try:
            # Abrir a caixa de diálogo para escolher o diretório e nome do arquivo
            filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if filepath:
                # Salvar o DataFrame em um arquivo Excel
                with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
                    resultado_concatenado.to_excel(writer, index=False)

                messagebox.showinfo('Exportação Concluída', f'Dados exportados com sucesso para:\n{filepath}')
        except Exception as e:
            messagebox.showerror('Erro ao Exportar', f'Ocorreu um erro ao exportar os dados:\n{str(e)}')
    else:
        messagebox.showwarning('Nenhum Dado', 'Não há dados para exportar.')

# Função para exportar resultados em PDF
def exportar_resultados_pdf():
    global resultado_concatenado

    if resultado_concatenado is not None and not resultado_concatenado.empty:
        try:
            # Abrir a caixa de diálogo para escolher o diretório e nome do arquivo
            filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])

            if filepath:
                output_dir = os.path.join(os.path.dirname(filepath), 'PDF')
                os.makedirs(output_dir, exist_ok=True)

                # Classe para gerar o PDF
                class PDF(FPDF):
                    def header(self):
                        if hasattr(self, 'logo_path'):
                            self.image(self.logo_path, 10, 8, 33)
                        self.set_font('Arial', 'B', 12)
                        self.cell(0, 10, 'Relatorio de Inconsistências', 0, 1, 'C')
                        self.ln(10)

                    def footer(self):
                        self.set_y(-30)
                        self.set_font('Arial', 'I', 8)
                        self.cell(0, 10, 'Fonte: Setor GIS NMC : Pedro Reis, Carolina Peres, Melquisedeque Nunes,Tiago Holanda', 0, 1, 'C')
                        self.set_y(-15)
                        self.cell(0, 10, 'Página %s' % self.page_no(), 0, 0, 'C')

                    def chapter_title(self, title):
                        self.set_font('Arial', 'B', 16)
                        self.cell(0, 10, title, 0, 1, 'C')
                        self.ln(10)

                    def chapter_body(self, body):
                        self.set_font('Arial', '', 12)
                        self.multi_cell(0, 10, body)
                        self.ln()

                    def table_row(self, row):
                        self.set_font('Arial', '', 12)
                        for item in row:
                            self.cell(0, 10, f'{item}', border=1, ln=0)
                        self.ln(10)

                logo_path = r"C:\Users\Tiago Holanda NMC\Desktop\NMC\dados pedro\MAC\Relatórios por WhatsApp\download-removebg-preview.png"  # Atualize com o caminho para o logotipo desejado

                # Gerando um PDF para cada linha
                for index, row in resultado_concatenado.iterrows():
                    pdf = PDF()
                    pdf.logo_path = logo_path
                    pdf.add_page()
                    pdf.chapter_title(f'Resultados para Linha {index + 1}')
                    
                    for column, value in row.items():
                        pdf.set_font('Arial', 'B', 12)
                        column_width = pdf.get_string_width(f'{column}:') + 10
                        pdf.cell(column_width, 10, f'{column}:', border=1)
                        pdf.set_font('Arial', '', 12)
                        pdf.multi_cell(0, 10, f'{value}', border=1)
                        pdf.ln(5)  
                    
                    output_path = os.path.join(output_dir, f'Linha_{index + 1}.pdf')
                    pdf.output(output_path)

                messagebox.showinfo('Exportação Concluída', 'PDFs gerados com sucesso!')
        except Exception as e:
            messagebox.showerror('Erro ao Exportar', f'Ocorreu um erro ao exportar os dados:\n{str(e)}')
    else:
        messagebox.showwarning('Nenhum Dado', 'Não há dados para exportar.')

# Função para iniciar a verificação de análise de preenchimento
def iniciar_analise_preenchimento():
    if 'df_bd' in globals() and 'df_matriz' in globals():
        resultado = AnalisePreenchimento(df_bd, df_matriz)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados e matriz antes de iniciar a análise.')

# Função para iniciar a verificação de análise de tipos
def iniciar_analise_tipos():
    if 'df_bd' in globals() and 'df_matriz' in globals():
        resultado = AnalisarTipos(df_bd, df_matriz)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados e matriz antes de iniciar a análise.')

# Função para iniciar a verificação de ortografia
def iniciar_verificacao_ortografia():
    if 'df_bd' in globals() and 'df_matriz' in globals():
        resultado = main_verificar_ortografia(df_matriz, df_bd)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados e matriz antes de iniciar a verificação.')

# Função para iniciar a verificação de nomes
def iniciar_verificacao_nomes():
    if 'df_bd' in globals() and 'df_matriz' in globals() and 'df_nomes' in globals():
        resultado = verificar_nomes(df_matriz, df_bd, df_nomes)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados, matriz e nomes antes de iniciar a verificação.')

# Função para iniciar a verificação de intervalos
def iniciar_verificacao_intervalos():
    if 'df_bd' in globals() and 'df_matriz' in globals() and 'df_intervalos' in globals():
        resultado = verificar_intervalos(df_matriz, df_bd, df_intervalos)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados, matriz e intervalos antes de iniciar a verificação.')

# Função para iniciar a verificação de condicionais
def iniciar_verificacao_condicionais():
    if 'df_bd' in globals() and 'df_matriz' in globals():
        resultado = verificar_condicionais(df_matriz, df_bd)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Adicionar os resultados ao widget Text
        text_resultado.insert(tk.END, resultado.to_string(index=False))
    else:
        messagebox.showwarning('Atenção', 'Selecione os arquivos de banco de dados e matriz antes de iniciar a verificação.')

# Função para concatenar todos os resultados em um relatório
def concatenar_resultados():
    global resultado_concatenado
    if 'df_bd' in globals() and 'df_matriz' in globals():
        resultado1 = AnalisePreenchimento(df_bd, df_matriz)
        resultado2 = AnalisarTipos(df_bd, df_matriz)
        resultado3 = main_verificar_ortografia(df_matriz, df_bd)
        resultado4 = verificar_nomes(df_matriz, df_bd, df_nomes) if 'df_nomes' in globals() else pd.DataFrame()
        resultado5 = verificar_intervalos(df_matriz, df_bd, df_intervalos) if 'df_intervalos' in globals() else pd.DataFrame()
        resultado6 = verificar_condicionais(df_matriz, df_bd)

        resultado_concatenado = pd.concat([resultado1, resultado2, resultado3, resultado4, resultado5, resultado6], ignore_index=True)

        # Limpar a área de resultados anteriores
        text_resultado.delete('1.0', tk.END)

        # Exibir as primeiras linhas do relatório
        text_resultado.insert(tk.END, resultado_concatenado.head(20).to_string(index=False))

    else:
        messagebox.showwarning('Atenção', 'Selecione todos os arquivos necessários antes de executar as verificações.')

# Função para selecionar arquivo de banco de dados
def selecionar_arquivo_banco_dados():
    caminho_banco_de_dados = filedialog.askopenfilename(title='Selecione o arquivo Excel do banco de dados')
    if caminho_banco_de_dados:
        try:
            global df_bd
            df_bd = pd.read_excel(caminho_banco_de_dados)
            btn_banco_de_dados.config(bg='#b5b5b5')
        except FileNotFoundError:
            messagebox.showerror('Erro', 'Arquivo do banco de dados não encontrado.')
        except pd.errors.ParserError:
            messagebox.showerror('Erro', 'Falha ao ler arquivo do banco de dados. Verifique se é um arquivo Excel válido.')
        except Exception as e:
            messagebox.showerror('Erro', f'Ocorreu um erro: {str(e)}')

# Função para selecionar arquivo da matriz
def selecionar_arquivo_matriz():
    caminho_matriz = filedialog.askopenfilename(title='Selecione o arquivo Excel da matriz')
    if caminho_matriz:
        try:
            global df_matriz
            df_matriz = pd.read_excel(caminho_matriz, header=1)
            btn_matriz.config(bg='#b5b5b5')
        except FileNotFoundError:
            messagebox.showerror('Erro', 'Arquivo da matriz não encontrado.')
        except pd.errors.ParserError:
            messagebox.showerror('Erro', 'Falha ao ler arquivo da matriz. Verifique se é um arquivo Excel válido.')
        except Exception as e:
            messagebox.showerror('Erro', f'Ocorreu um erro: {str(e)}')

# Função para selecionar arquivo de nomes
def selecionar_arquivo_nomes():
    caminho_nomes = filedialog.askopenfilename(title='Selecione o arquivo Excel de nomes')
    if caminho_nomes:
        try:
            global df_nomes
            df_nomes = pd.read_excel(caminho_nomes)
            btn_nomes.config(bg='#b5b5b5')
        except FileNotFoundError:
            messagebox.showerror('Erro', 'Arquivo de nomes não encontrado.')
        except pd.errors.ParserError:
            messagebox.showerror('Erro', 'Falha ao ler arquivo de nomes. Verifique se é um arquivo Excel válido.')
        except Exception as e:
            messagebox.showerror('Erro', f'Ocorreu um erro: {str(e)}')

# Função para selecionar arquivo de intervalos
def selecionar_arquivo_intervalos():
    caminho_intervalos = filedialog.askopenfilename(title='Selecione o arquivo Excel de intervalos')
    if caminho_intervalos:
        try:
            global df_intervalos
            df_intervalos = pd.read_excel(caminho_intervalos)
            btn_intervalos.config(bg='#b5b5b5')
        except FileNotFoundError:
            messagebox.showerror('Erro', 'Arquivo de intervalos não encontrado.')
        except pd.errors.ParserError:
            messagebox.showerror('Erro', 'Falha ao ler arquivo de intervalos. Verifique se é um arquivo Excel válido.')
        except Exception as e:
            messagebox.showerror('Erro', f'Ocorreu um erro: {str(e)}')

# Configuração da interface gráfica
root = tk.Tk()
root.title('Verificador de Inconsistências - NMC')
root.configure(bg='#252859')
root.resizable(False,False)

root.iconbitmap(r'C:\\Users\\Tiago Holanda NMC\\Desktop\\NMC\\dados pedro\\MAC\\download-removebg-preview.png')

def conf_wid(widget):
    estilo_fonte = font.Font(size=9, weight='bold')
    widget.configure(fg='white', font=estilo_fonte)

# Frames
frame_importacao = tk.LabelFrame(root, text='Importar Arquivos', padx=10, pady=10)
frame_importacao.grid(row=0, column=0, padx=10, pady=10, sticky='ew')
frame_importacao.configure(bg='#252859')
conf_wid(frame_importacao)

frame_analises = tk.LabelFrame(root, text='Executar Análises', padx=10, pady=10)
frame_analises.grid(row=1, column=0, padx=10, pady=10, sticky='ew')
frame_analises.configure(bg='#252859')
conf_wid(frame_analises)

frame_resultado = tk.LabelFrame(root, text='Resultados das Análises', padx=10, pady=10)
frame_resultado.grid(row=2, column=0, padx=10, pady=10, sticky='nsew')
frame_resultado.configure(bg='#252859')
conf_wid(frame_resultado)

# Botões de importação
btn_banco_de_dados = tk.Button(frame_importacao, text='Selecionar Banco de Dados', command=selecionar_arquivo_banco_dados, width=20)
btn_banco_de_dados.grid(row=0, column=0, padx=5, pady=5)

btn_matriz = tk.Button(frame_importacao, text='Selecionar Matriz', command=selecionar_arquivo_matriz,width=20)
btn_matriz.grid(row=0, column=1, padx=5, pady=5)

btn_intervalos = tk.Button(frame_importacao, text='Selecionar Intervalos', command=selecionar_arquivo_intervalos,width=20)
btn_intervalos.grid(row=1, column=0, padx=5, pady=5)

btn_nomes = tk.Button(frame_importacao, text='Selecionar Nomes', command=selecionar_arquivo_nomes,width=20)
btn_nomes.grid(row=1, column=1, padx=5, pady=5)

# Botões para executar análises individuais
btn_analise_preenchimento = tk.Button(frame_analises, text='Análise de Preenchimento', command=iniciar_analise_preenchimento,width=20)
btn_analise_preenchimento.grid(row=0, column=4, padx=5, pady=5)

btn_analise_tipos = tk.Button(frame_analises, text='Análise de Tipos', command=iniciar_analise_tipos,width=20)
btn_analise_tipos.grid(row=1, column=4, padx=5, pady=5)

btn_verificacao_ortografia = tk.Button(frame_analises, text='Verificar Ortografia', command=iniciar_verificacao_ortografia,width=20)
btn_verificacao_ortografia.grid(row=0, column=2, padx=5, pady=5)

btn_verificacao_nomes = tk.Button(frame_analises, text='Verificar Nomes', command=iniciar_verificacao_nomes,width=20)
btn_verificacao_nomes.grid(row=1, column=2, padx=5, pady=5)

btn_verificacao_intervalos = tk.Button(frame_analises, text='Verificar Intervalos', command=iniciar_verificacao_intervalos,width=20)
btn_verificacao_intervalos.grid(row=0, column=3, padx=5, pady=5)

btn_verificacao_condicionais = tk.Button(frame_analises, text='Verificar Condicional', command=iniciar_verificacao_condicionais,width=20)
btn_verificacao_condicionais.grid(row=1, column=3, padx=5, pady=5)

# Botão para concatenar resultados em relatório
btn_concatenar_resultados = tk.Button(frame_analises, text='Executar Verificações', command=concatenar_resultados,width=20)
btn_concatenar_resultados.grid(row=0, column=0, padx=(5,180), pady=5)

# Botão para exportar resultados
btn_exportar_resultados = tk.Button(frame_analises, text='Exportar Resultados', command=exportar_resultados_excel, width=20)
btn_exportar_resultados.grid(row=1, column=0, padx=(5,180), pady=5)

# Botão para exportar resultados em PDF
btn_exportar_resultados_pdf = tk.Button(frame_analises, text='Exportar Resultados em PDF', command=exportar_resultados_pdf, width=20)
btn_exportar_resultados_pdf.grid(row=1, column=1, padx=5, pady=5)

# Área de resultados das análises
text_resultado = tk.Text(frame_resultado, height=20, width=100)
text_resultado.grid(row=0, column=0, padx=5, pady=5)
scrollbar = ttk.Scrollbar(frame_resultado, orient=tk.VERTICAL, command=text_resultado.yview)
scrollbar.grid(row=0, column=1, sticky='ns')
text_resultado.config(yscrollcommand=scrollbar.set)

# Configuração de redimensionamento das células
root.columnconfigure(0, weight=1)
root.rowconfigure(2, weight=1)

# Loop principal da aplicação
root.mainloop()

print("PDFs gerados com sucesso!")
